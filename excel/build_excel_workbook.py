"""
build_excel_workbook.py
Builds Retail_Analytics_Workbook.xlsx with:
  - Raw_Segments (data)
  - Raw_Monthly_Revenue (data)
  - Raw_Churn_Risk (data, top 100)
  - Exec_Summary (SUMIFS/COUNTIFS/AVERAGEIFS formulas referencing raw tabs)
  - Promo_Scenario (sensitivity / what-if formulas, Solver-ready structure)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
LABEL_FONT = Font(name=FONT, bold=True, size=10)
BODY_FONT = Font(name=FONT, size=10)
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

segments = pd.read_csv("../data/customer_segments.csv")
monthly = pd.read_csv("../data/monthly_revenue.csv")
churn = pd.read_csv("../data/churn_predictions.csv").sort_values("churn_probability", ascending=False).head(100)
ab = pd.read_csv("../data/ab_test_results.csv")

wb = Workbook()

# ---------------------------------------------------------------
# Sheet: Raw_Segments
# ---------------------------------------------------------------
ws = wb.active
ws.title = "Raw_Segments"
cols = ["customer_id", "recency_days", "frequency", "monetary", "rfm_segment", "segment_label"]
ws.append(cols)
for c in ws[1]:
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for _, row in segments[cols].iterrows():
    ws.append(list(row))
for col_cells in ws.columns:
    length = max(len(str(c.value)) for c in col_cells)
    ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 22)
ws.freeze_panes = "A2"

# ---------------------------------------------------------------
# Sheet: Raw_Monthly_Revenue
# ---------------------------------------------------------------
ws2 = wb.create_sheet("Raw_Monthly_Revenue")
cols2 = ["month", "revenue", "prev_month_revenue", "mom_growth_pct", "revenue_12mo_ago", "yoy_growth_pct"]
ws2.append(cols2)
for c in ws2[1]:
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for _, row in monthly[cols2].iterrows():
    ws2.append(list(row))
for col_cells in ws2.columns:
    ws2.column_dimensions[col_cells[0].column_letter].width = 16
ws2.freeze_panes = "A2"

# ---------------------------------------------------------------
# Sheet: Raw_Churn_Risk (top 100 at-risk customers)
# ---------------------------------------------------------------
ws3 = wb.create_sheet("Raw_Churn_Risk")
cols3 = ["customer_id", "churn_probability", "orders_0_180", "spend_0_180"]
ws3.append(cols3)
for c in ws3[1]:
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for _, row in churn[cols3].iterrows():
    ws3.append(list(row))
for col_cells in ws3.columns:
    ws3.column_dimensions[col_cells[0].column_letter].width = 18
ws3.freeze_panes = "A2"

# ---------------------------------------------------------------
# Sheet: Exec_Summary  (formulas reference the raw tabs)
# ---------------------------------------------------------------
es = wb.create_sheet("Exec_Summary", 0)
es.sheet_view.showGridLines = False
es["B2"] = "Retail & E-Commerce Customer Intelligence — Executive Summary"
es["B2"].font = TITLE_FONT
es.merge_cells("B2:G2")

n_seg = len(segments)
n_month = len(monthly)

kpi_labels = ["Total Customers (RFM base)", "Total Revenue (all-time)", "Avg Order Value proxy (Monetary/Frequency)",
              "Latest Month Revenue", "Latest Month MoM Growth %", "Latest Month YoY Growth %"]
es["B4"] = "Key Performance Indicators"
es["B4"].font = LABEL_FONT
row0 = 5
kpi_formulas = [
    f"=COUNTA(Raw_Segments!A2:A{n_seg+1})",
    f"=SUM(Raw_Segments!D2:D{n_seg+1})",
    f"=ROUND(SUM(Raw_Segments!D2:D{n_seg+1})/SUM(Raw_Segments!C2:C{n_seg+1}),2)",
    f"=INDEX(Raw_Monthly_Revenue!B2:B{n_month+1},{n_month})",
    f"=INDEX(Raw_Monthly_Revenue!D2:D{n_month+1},{n_month})",
    f"=INDEX(Raw_Monthly_Revenue!F2:F{n_month+1},{n_month})",
]
for i, (label, formula) in enumerate(zip(kpi_labels, kpi_formulas)):
    r = row0 + i
    es[f"B{r}"] = label
    es[f"B{r}"].font = BODY_FONT
    es[f"D{r}"] = formula
    es[f"D{r}"].font = Font(name=FONT, bold=True, size=11)
    if "%" in label:
        es[f"D{r}"].number_format = '0.0"%"'
    elif "Revenue" in label or "Order Value" in label:
        es[f"D{r}"].number_format = "$#,##0"

# Segment breakdown table (COUNTIFS / SUMIFS / AVERAGEIFS)
es["B13"] = "Segment Breakdown (RFM K-Means Clusters)"
es["B13"].font = LABEL_FONT
headers = ["Segment", "Customers", "% of Base", "Total Revenue", "Avg Monetary Value", "Avg Recency (days)"]
for j, h in enumerate(headers):
    cell = es.cell(row=14, column=2 + j, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

unique_segments = segments["segment_label"].unique().tolist()
for i, seg_name in enumerate(unique_segments):
    r = 15 + i
    es.cell(row=r, column=2, value=seg_name).font = BODY_FONT
    es.cell(row=r, column=3, value=f'=COUNTIF(Raw_Segments!$F$2:$F${n_seg+1},B{r})')
    es.cell(row=r, column=4, value=f'=ROUND(C{r}/$D$5,3)')
    es.cell(row=r, column=4).number_format = "0.0%"
    es.cell(row=r, column=5, value=f'=SUMIF(Raw_Segments!$F$2:$F${n_seg+1},B{r},Raw_Segments!$D$2:$D${n_seg+1})')
    es.cell(row=r, column=5).number_format = "$#,##0"
    es.cell(row=r, column=6, value=f'=ROUND(AVERAGEIF(Raw_Segments!$F$2:$F${n_seg+1},B{r},Raw_Segments!$D$2:$D${n_seg+1}),0)')
    es.cell(row=r, column=6).number_format = "$#,##0"
    es.cell(row=r, column=7, value=f'=ROUND(AVERAGEIF(Raw_Segments!$F$2:$F${n_seg+1},B{r},Raw_Segments!$B$2:$B${n_seg+1}),1)')
    for col in range(2, 8):
        es.cell(row=r, column=col).border = BORDER

# column widths
for col, w in zip("BCDEFG", [30, 14, 12, 16, 18, 16]):
    es.column_dimensions[col].width = w

# Bar chart: revenue by segment
chart = BarChart()
chart.title = "Total Revenue by Segment"
chart.y_axis.title = "Revenue ($)"
data_ref = Reference(es, min_col=5, min_row=14, max_row=14 + len(unique_segments))
cats_ref = Reference(es, min_col=2, min_row=15, max_row=14 + len(unique_segments))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 16
chart.height = 8
es.add_chart(chart, "I14")

# Line chart: monthly revenue trend
line = LineChart()
line.title = "Monthly Revenue Trend"
line.y_axis.title = "Revenue ($)"
rev_ref = Reference(ws2, min_col=2, min_row=1, max_row=n_month + 1)
month_ref = Reference(ws2, min_col=1, min_row=2, max_row=n_month + 1)
line.add_data(rev_ref, titles_from_data=True)
line.set_categories(month_ref)
line.width = 16
line.height = 8
es.add_chart(line, "I32")

# ---------------------------------------------------------------
# Sheet: Promo_Scenario (what-if sensitivity, Solver-ready)
# ---------------------------------------------------------------
ps = wb.create_sheet("Promo_Scenario")
ps.sheet_view.showGridLines = False
ps["B2"] = "Promo Discount — Revenue & Margin Sensitivity Model"
ps["B2"].font = TITLE_FONT
ps.merge_cells("B2:F2")

ps["B4"] = "Assumptions (edit the yellow cells)"
ps["B4"].font = LABEL_FONT

assumptions = [
    ("Baseline customers reached", 7580, None),
    ("Baseline conversion rate (no discount)", 0.708, "0.0%"),
    ("Baseline Average Order Value ($)", 1131, "$#,##0"),
    ("Conversion lift per 1% discount (elasticity assumption)", 0.004, "0.0%"),
    ("Avg product margin (% of price, before discount)", 0.42, "0.0%"),
]
for i, (label, val, fmt) in enumerate(assumptions):
    r = 5 + i
    ps.cell(row=r, column=2, value=label).font = BODY_FONT
    c = ps.cell(row=r, column=5, value=val)
    c.font = INPUT_FONT
    c.fill = YELLOW_FILL
    if fmt:
        c.number_format = fmt

ps["B12"] = "Discount Scenario Table"
ps["B12"].font = LABEL_FONT
headers2 = ["Discount %", "Projected Conversion Rate", "Customers Converted",
            "Effective AOV ($)", "Projected Revenue ($)", "Projected Gross Margin ($)"]
for j, h in enumerate(headers2):
    cell = ps.cell(row=13, column=2 + j, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

discounts = [0.00, 0.05, 0.10, 0.15, 0.20, 0.25, 0.30]
for i, d in enumerate(discounts):
    r = 14 + i
    ps.cell(row=r, column=2, value=d).number_format = "0%"
    ps.cell(row=r, column=2).font = BODY_FONT
    # conversion rate = baseline + elasticity * discount(in whole points)
    ps.cell(row=r, column=3, value=f"=$E$6+($E$8*B{r}*100)")
    ps.cell(row=r, column=3).number_format = "0.0%"
    # customers converted = baseline reached * conv rate
    ps.cell(row=r, column=4, value=f"=ROUND($E$5*C{r},0)")
    # effective AOV = baseline AOV * (1 - discount)
    ps.cell(row=r, column=5, value=f"=$E$7*(1-B{r})")
    ps.cell(row=r, column=5).number_format = "$#,##0"
    # projected revenue = customers converted * effective AOV
    ps.cell(row=r, column=6, value=f"=D{r}*E{r}")
    ps.cell(row=r, column=6).number_format = "$#,##0"
    # projected gross margin = revenue * (margin% adjusted for discount eating into margin first)
    ps.cell(row=r, column=7, value=f"=F{r}*MAX($E$9-B{r},0)")
    ps.cell(row=r, column=7).number_format = "$#,##0"
    for col in range(2, 8):
        ps.cell(row=r, column=col).border = BORDER

ps["B23"] = "Note: this scenario grid uses live formulas, so it updates instantly when you change the yellow assumption cells."
ps["B23"].font = Font(name=FONT, italic=True, size=9, color="666666")
ps.merge_cells("B23:G23")
ps["B25"] = ("To find the margin-maximizing discount automatically: Data > What-If Analysis > Solver. "
             "Set Objective = the max of column G, By Changing Cell = a single discount input cell, "
             "Constraint: 0% <= discount <= 30%. Solver is a native Excel add-in and must be run "
             "inside Excel/LibreOffice — it can't be pre-baked into a file.")
ps["B25"].font = Font(name=FONT, italic=True, size=9, color="666666")
ps.merge_cells("B25:G26")

for col, w in zip("BCDEFG", [20, 22, 20, 16, 18, 20]):
    ps.column_dimensions[col].width = w

wb.save("../excel/Retail_Analytics_Workbook.xlsx")
print("Saved Retail_Analytics_Workbook.xlsx")
